import requests
import selenium
import time
import pdb
from selenium import webdriver
import logging
import re
import datetime
import os
import openpyxl
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup as b
from selenium.webdriver.support.select import Select
class instabot:
    def __init__(self, driver, username, pas):
        self.user = username
        self.password = pas
        self.driver = driver
        self.date_obj = datetime.date.today()

    def opening_files_for_data_saving(self, name):
        parent_dir = 'D:/'
        directory =  self.user
        dirr = os.path.join(parent_dir, directory+'/')
        if os.path.exists(dirr):
            print('path exist')
        else:
            print("making a new directory: "+ dirr)
            os.mkdir(dirr)
        if os.path.exists(dirr+ self.user + '.xlsx'):
            print('opening the xlsx file: '+ dirr + '/' + self.user +'.xlsx')
            wb = openpyxl.load_workbook(dirr + '/' + self.user + '.xlsx')
        else:
            path = os.getcwd()
            print(path)
            print('creating a new xlsx file for the data base: '+ dirr + '/' + self.user +'.xlsx')
            wb = openpyxl.Workbook()
        try:
            sheet = wb[name]
            print('Sheet ' +name+' exists opening it.')
        except:
            print('Sheet ' +name+' does not exists creating it.')
            sheet = wb.create_sheet(name)
        return dirr, wb, sheet

    def login(self):
        try:
            self.driver.get('https://www.instagram.com/')
        except:
            logging.error('failed to load please check internet connection or the url, if it is correct?')
        time.sleep(2)
        self.driver.find_element_by_xpath('//input[contains(@aria-label,"username")]').send_keys(self.user)
        self.driver.find_element_by_xpath('//input[contains(@aria-label,"Password")]').send_keys(self.password)
        self.driver.find_element_by_xpath('//div[contains(text(), "Log In" )]').click()
        time.sleep(4)
        self.driver.find_element_by_xpath('//button[contains(text(), "Not Now" )]').click()
        time.sleep(4)
        self.driver.find_element_by_xpath('//button[contains(text(), "Not Now")]').click()
        #self.opening_files_for_data_saving(self.user)

    def fetching_followings_count_info(self):
        time.sleep(3)
        #self.driver.find_element_by_xpath('//div[@class="RR-M-  _2NjG_"]//a[@class="_2dbep qNELH kIKUG"]').click()
        jam = self.driver.find_element_by_xpath('//div[@class="Fifk5"]//span//img')
        jam.click()
        select = self.driver.find_element_by_xpath('//div[contains(text(), "Profile")]')
        select.click()
        time.sleep(3)
        self.driver.find_element_by_xpath('//a[contains(@href, "following")]').click()
        time.sleep(2)
        self.driver.execute_script("window.scrollBy(0,document.body.scrollHeight);")
        #last_height = self.driver.execute_script("return document.documentElement.scrollHeight")
        last_ht, ht =0, 1
        time.sleep(2)
        scrollbox = self.driver.find_element_by_xpath('/html/body/div[4]/div/div/div[2]')

        while last_ht != ht:
            last_ht = ht
            # Scroll down to bottom
            #self.driver.execute_script("window.scrollTo(0,document.documentElement.scrollHeight);")
            #self.driver.execute_scrript("arguments[0].scrollIntoView();",element)
            # Wait to load page
            ht = self.driver.execute_script("""
                    arguments[0].scrollTo(0, arguments[0].scrollHeight);
                    return arguments[0].scrollHeight;                        
            """, scrollbox)
            time.sleep(3)
        links = scrollbox.find_elements_by_tag_name('a')
        names = [name.text for name in links if name.text != '']
        names = set(names)
        print("***********you are following these many people = {}***************".format(len(names)))
        count = len(names)
        self.driver.find_element_by_xpath('//div[@class="WaOAr"]//button[@class="wpO6b "]//div').click()
        following_count = self.get_following_count()
        if float(count) >= following_count:
            print("saving the following list in an excel file")
            # opening the xlsx(excel) file where the following names is needed to be saved
            directoy, wb_obj, sheet = self.opening_files_for_data_saving('Following_names')
            lenn = len(sheet['A'])
            sheet['A' + str(lenn + 1)] = '***NAME OF THE USERS YOU ARE FOLLOWING on date' + str(self.date_obj) + '***'
            sheet.column_dimensions['A'].width = 70
            for name, i in zip(names, range(lenn+2, lenn+count+1)):
                sheet['A'+str(i)] = name
            wb_obj.save(directoy+self.user+'.xlsx')
            return count, names
        else:
            print('ALL the following name did not got loaded properly loading again')
            return self.fetching_followings_count_info()

    def fetching_followers_count_info(self):
        time.sleep(3)
        #self.driver.find_element_by_xpath('//div[@class="RR-M-  _2NjG_"]//a[@class="_2dbep qNELH kIKUG"]').click()
        jam = self.driver.find_element_by_xpath('//div[@class="Fifk5"]//span//img')
        jam.click()
        select = self.driver.find_element_by_xpath('//div[contains(text(), "Profile")]')
        select.click()
        time.sleep(3)
        self.driver.find_element_by_xpath('//ul[@class="k9GMp "]//a[contains(@href, "followers")]').click()
        time.sleep(2)
        self.driver.execute_script("window.scrollBy(0,document.body.scrollHeight);")
        last_ht, ht =0, 1
        time.sleep(1)
        scrollbox = self.driver.find_element_by_xpath('/html/body/div[4]/div/div/div[2]')
        while last_ht != ht:
            last_ht = ht
            # Scroll down to bottom
            ht = self.driver.execute_script("""
                    arguments[0].scrollTo(0, arguments[0].scrollHeight);
                    return arguments[0].scrollHeight;                        
            """, scrollbox )
            # Wait to load page
            time.sleep(2)
        links = scrollbox.find_elements_by_tag_name('a')
        names = [name.text for name in links if name.text != '']
        names = set(names)
        count = len(names)
        print("***********YOUR FOLLOWERS COUNT = {}***************".format(len(names)))
        self.driver.find_element_by_xpath('//div[@class="WaOAr"]//button[@class="wpO6b "]//div').click()
        no_of_followers = self.get_followers_count()
        if count >= no_of_followers:
            print("saving the followers list in an excel file")
            # opening the xlsx(excel) file where the followers names is needed to be saved
            directoy, wb_obj, sheet = self.opening_files_for_data_saving('Follwers_names')
            lenn = len(sheet['A'])
            sheet['A' + str(lenn + 1)] = '***NAME OF UR FOLLOWERS on date' + str(self.date_obj) + '***'
            sheet.column_dimensions['A'].width = 70
            for name, i in zip(names, range(lenn+2, lenn+count+1)):
                sheet['A'+str(i)] = name
            wb_obj.save(directoy+self.user+'.xlsx')
            return count, names
        else:
            print('ALL the followers name did not got loaded properly loading again')
            return self.fetching_followers_count_info()

    def fetching_greedy_people(self, following_names, followers_names):
        cheap_ppl = []
        for name in following_names:
            if name not in followers_names:
                cheap_ppl.append(name)
        cheap_ppl_count = len(cheap_ppl)
        print('**********************People count who does not deserve to follow - {}***********************'.format(cheap_ppl_count))
        print('*****************THE PEOPLE WHOME YOU NEED TO TAKE ACTION FOR **************')
        print("saving the greedy people list  in an excel file")
        directoy, wb_obj, sheet = self.opening_files_for_data_saving('people not following you')
        lenn = len(sheet['A'])
        sheet['A' + str(lenn + 1)] = '**THE PEOPLE WHOME YOU NEED TO TAKE ACTION FOR' + str(self.date_obj) + '**'
        for name, i in zip(cheap_ppl, range(lenn+2, lenn+cheap_ppl_count+1)):
            sheet['A'+str(i)] = name
        wb_obj.save(directoy + self.user + '.xlsx')
        return cheap_ppl_count, cheap_ppl

    def action_on_greedy_people(self, cheap_ppl):
       # directoy, wb_obj, sheet = self.opening_files_for_data_saving(self.user)
        unfollowing_ppl = []
        directoy, wb_obj, sheet_unfollowed = self.opening_files_for_data_saving('unfollowed_name')
        directoy, wb_obj, sheet_not_unfollowed = self.opening_files_for_data_saving('unfollowed_name')
        lenn_unfolled = len(sheet_unfollowed['A'])
        sheet_unfollowed['A'+str(lenn_unfolled+1)] = '**Peaple you have unfollowed on date'+str(self.date_obj)+'**'
        lenn_unfolled += 1
        lenn_not_unfollowed = len(sheet_not_unfollowed['A'])
        sheet_not_unfollowed['A' + str(lenn_not_unfollowed + 1)] = '**Peaple you did not unfollowed on date'+str(self.date_obj)+'**'
        lenn_not_unfollowed += 1
        for k in cheap_ppl:
            time.sleep(2)
            self.driver.get('https://www.instagram.com/{}/'.format(k))
            try:
                self.driver.find_element_by_xpath('//div[contains(@class,"error-container" )]')
                logging.warning('Loading error for user {} '.format(k))
                self.driver.login()
                continue
            except:
                logging.info('loaded successfully moving forward')

            time.sleep(3)
            total_post = self.get_no_posts()
            time.sleep(2)
            if float(total_post) < 600:
                logging.info('!!!!!!!!!!!!!!!!unfollowing : {}'.format(k))
                self.driver.find_element_by_xpath('//*[@class="vBF20 _1OSdk"]//button | //button[contains(@class,"L3NKy    _8A5w5 ")]').click()
                self.driver.find_element_by_xpath('//button[contains(text(), "Unfollow")]').click()
                unfollowing_ppl.append(k)
                #print("saving unfollowed people name in excel ")
                sheet_unfollowed['A' + str(lenn_unfolled+1)] = k + '- no of post :'+ str(total_post)
            else:
                #print("saving people name whome  you did not unfollowed in excel ")
                sheet_not_unfollowed['A' + str(lenn_not_unfollowed+1)] = k + '- no of post :'+str(total_post)

        wb_obj.save(directoy+self.user+'.xlsx')


    def get_followers_count(self):
        #self.driver.find_element_by_xpath('//div[@class="RR-M-  _2NjG_"]//a[@class="_2dbep qNELH kIKUG"]').click()
        #self.driver.get('https://www.instagram.com/{}/'.format(self.user))
        time.sleep(2)
        flw = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '#react-root > section > main')))
        sflw = b(flw.get_attribute('innerHTML'), 'html.parser')
        followers = sflw.findAll('span', {'class': 'g47SY'})
        f = followers[1].getText().replace(',', '')
        if 'k' in f:
            f = float(f[:-1]) * 10 ** 3
            return f
        elif 'm' in f:
            f = float(f[:-1]) * 10 ** 6
            return f
        else:
            return float(f)

    def get_following_count(self):
        #self.driver.find_element_by_xpath('//div[@class="RR-M-  _2NjG_"]//a[@class="_2dbep qNELH kIKUG"]').click()
        #self.driver.get('https://www.instagram.com/{}/'.format(self.user))
        time.sleep(2)
        flw = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '#react-root > section > main')))
        sflw = b(flw.get_attribute('innerHTML'), 'html.parser')
        followers = sflw.findAll('span', {'class': 'g47SY'})
        f = followers[2].getText().replace(',', '')
        if 'k' in f:
            f = float(f[:-1]) * 10 ** 3
            return f
        elif 'm' in f:
            f = float(f[:-1]) * 10 ** 6
            return f
        else:
            return float(f)

    def get_no_posts(self):
        #self.driver.find_element_by_xpath('//div[@class="RR-M-  _2NjG_"]//a[@class="_2dbep qNELH kIKUG"]').click()
        #self.driver.get('https://www.instagram.com/{}/'.format(self.user))
        time.sleep(2)
        # jam = self.driver.find_element_by_xpath('//div[@class="Fifk5"]//span')
        # jam.click()
        # select = self.driver.find_element_by_xpath('//div[contains(text(), "Profile")]')
        # select.click()
        time.sleep(1)
        flw = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '#react-root > section > main')))
        sflw = b(flw.get_attribute('innerHTML'), 'html.parser')
        followers = sflw.findAll('span', {'class': 'g47SY'})
        f = followers[0].getText().replace(',', '')
        if 'k' in f:
            f = float(f[:-1]) * 10 ** 3
            return f
        elif 'm' in f:
            f = float(f[:-1]) * 10 ** 6
            return f
        else:
            return float(f)

    def returns_followers_names(self, fetch_for):
        self.driver.get('https://www.instagram.com/{}/'.format(fetch_for))
        time.sleep(3)
        pdb.set_trace()
        #self.driver.execute_script("window.scrollBy(0,200)")
        self.driver.execute_script("window.scrollBy(0,-10000)")
        self.driver.find_element_by_css_selector('#react-root > section > main > div > header > section > ul > li:nth-child(2) > a').click()
        time.sleep(2)
        self.driver.execute_script("window.scrollBy(0,document.body.scrollHeight);")
        last_ht, ht =0, 1
        time.sleep(1)
        scrollbox = self.driver.find_element_by_xpath('/html/body/div[4]/div/div/div[2]')
        while last_ht != ht:
            last_ht = ht
            # Scroll down to bottom
            ht = self.driver.execute_script("""
                    arguments[0].scrollTo(0, arguments[0].scrollHeight);
                    return arguments[0].scrollHeight;                        
            """, scrollbox )
            # Wait to load page
            time.sleep(2)
        links = scrollbox.find_elements_by_tag_name('a')
        names = [name.text for name in links if name.text != '']
        names = set(names)
        count = len(names)
        self.driver.find_element_by_xpath('//div[@class="WaOAr"]//button[@class="wpO6b "]//div').click()
        no_of_followers = self.get_followers_count()
        print("************count =" +str(count)+"*********")
        if count + 10 >= no_of_followers:
            return count, names
        else:
            print('ALL the followers name did not got loaded properly loading again')
            return self.returns_followers_names(fetch_for)

    def stale_send_follow_request(self,fetch_for):
        self.driver.get('https://www.instagram.com/{}/'.format(fetch_for))
        time.sleep(3)
        if self.is_public():
            self.driver.find_element_by_xpath('//ul[@class="k9GMp "]//a[contains(@href, "followers")]').click()
            time.sleep(2)
            self.driver.execute_script("window.scrollBy(0,document.body.scrollHeight);")
            last_ht, ht = 0, 1
            time.sleep(1)
            scrollbox = self.driver.find_element_by_xpath('/html/body/div[4]/div/div/div[2]')
            while last_ht != ht:
                last_ht = ht
                # Scroll down to bottom
                ht = self.driver.execute_script("""
                         arguments[0].scrollTo(0, arguments[0].scrollHeight);
                         return arguments[0].scrollHeight;                        
                 """, scrollbox)
                # Wait to load page
                time.sleep(2)
            links = scrollbox.find_elements_by_tag_name('button')
            for button in links:
                if button.text == 'Following' or  button.text == 'Requested':
                    continue
                else:
                    try:
                        button.click()
                        time.sleep(3)
                    except:
                        print('enable to follow moving forward')


    def is_public(self):
        try:
            astate = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.CLASS_NAME, 'rkEop')))
            if astate.text == 'This Account is Private':
                return False
            else:
                return True
        except:
            return True


    def like_post(self):
        post = self.driver.find_element_by_css_selector(
            '#react-root > section > main > div > div._2z6nI > article > div > div > div:nth-child(1) > div:nth-child(1)')
        html = post.get_attribute('innerHTML')
        h = b(html, 'html.parser')
        href = h.a['href']
        self.driver.get('https://www.instagram.com' + href)
        like_btn = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR,'.fr66n > button')))
        like_btn.click()


    def follow_page(self):
        try:
            follow = WebDriverWait(self.driver, 3).until(EC.presence_of_element_located(
                (By.XPATH, '//button[contains(text(), "Follow")]')))
            follow.click()
        except:
            print('already following')
            raise Exception
            time.sleep(1)

#obj = instabot('', '')
#obj.login('', '')
#print(obj.get_followers_count())
#print(obj.get_following_count())
# following_count, following_names = obj.fetching_followings_count_info()
# followers_count, followers_names = obj.fetching_followers_count_info()
# cheap_ppl_count, cheap_ppl = obj.fetching_greedy_people(following_names, followers_names)
# obj.action_on_greedy_people(cheap_ppl)

#obj.send_follow_request('seju_craft_industry_16')